import random

from initialization.utils.llm_util import request_response
import re
import logging
import openpyxl

logger = logging.getLogger(__name__)


class Extractor:
    def __init__(self):
        self.role_prompt = ("You are an expert in Simulation-based Testing for Autonomous Driving Systems, "
                            "with the goal of extracting functional scenarios from public accident reports. "
                            "Here is a description of an accident report: ")

        # self.task_summarize_prompt = ("Please identify the actions of each vehicle and summarize these actions as a short"
        #                               "description of the functional scenario. ")
        self.task_summarize_prompt = ("Please list the \'Road Structure\', \'Initial Actions\' of vehicles and the \'Interactive Pattern Sequence\' between pair of vehicles. "
                                      "The road structure should be like \"straight 4-lane road\", the initial actions should be like \" (V1): V1 drives on ...; ...\". "
                                      "Each interactive pattern should only contain 2 vehicles and follow the format \"(Vi, Vj): action description of Vi and Vj. \".")
        # self.task_extract_param_prompt = "Then extract the parameter ranges of vehicle actions from the report, " \
        #                                  "and list them below without other notes. "
        # self.task_select_ego_prompt = "In this scenario, which vehicle performs less actions but faces a complex and dangerous scenario? "
        self.attention_summarize_prompt = (
            "Attention: \n"
            "1. the action description should contain the relative road position of each vehicle and be in one sentence; \n"
            "2. the verb to describe each action should be selected from {brake, decelerate, accelerate, change lane left/right, left/right/U turn (intersection)};\n"
            "3. do not involve the interactive pattern after the first crash occurs;\n"
            "4. focus on the vehicles' movement and do not describe the drivers' actions.\n"
            "5. try to make the interactive pattern sequence as short as possible.\n"
           )
        # "2. the verbs that describe each action is selected from {brake, decelerate, accelerate, swerve left/right, lane change left/right}; "?
        # "3. do not involve the continue driving/collide/struck/rest of vehicles in the pattern sequence;"
        # self.attention_select_ego_prompt = "Attention: just list the vehicle name without other information. "

    @staticmethod
    def wrap_user_message(message):
        return {"role": "user", "content": message}

    @staticmethod
    def wrap_system_message(message):
        return {"role": "assistant", "content": message}

    def extract(self, report):
        dialogue_history = [{"role": "system", "content": self.role_prompt}]

        message1 = report + "\n" + self.task_summarize_prompt + "\n" + self.attention_summarize_prompt
        print("------------- Message Current --------------------\n", message1)
        dialogue_history.append(self.wrap_user_message(message1))
        response = request_response(dialogue_history, task_id=1)
        # print("DEBUG response:", response)
        response = response.choices[0].message.content
        print("model response: \n", response)
        logger.info("Model Response: \n %s", response)
        dialogue_history.append(self.wrap_system_message(response))
        func_scenario, func_scenario_dict, candidate_ego = self.process_response(response)
        logger.info("Extracted Functional Scenario: \n %s \n %s", func_scenario, str(candidate_ego))

        extracted_data = {"func_scenario": func_scenario, "func_scenario_dict": func_scenario_dict, "candidate_ego": random.choice(candidate_ego)}
        logger.info("Extracted Data: %s", extracted_data)
        # print(extracted_data)

        return extracted_data

    @staticmethod
    def process_response(response):
        # func_scenario = re.search(r"^(.*?)(?=parameter ranges:|\n)",
        #                           response,
        #                           re.IGNORECASE | re.DOTALL
        #                           ).group(0).strip()
        # initial_string = re.search(r"\((.*?)\):\s(.*?)\.", response).group(0)
        road_structure = re.search(r'(Road Structure.*)', response).group(1)
        print(road_structure[road_structure.find(':') + 1:])
        response = response.replace(road_structure, "")
        initial_string = re.search(r'(?i):(.*?)Interactive Pattern', response, re.DOTALL).group(1)
        response = response.replace(initial_string, "")
        # vehicle_list = initial_string.split(':')[0].strip()[1:-1].split(", ")
        vehicle_list = re.findall(r'V\d+', response)
        frequency_dict = {}
        for v in set(vehicle_list):
            frequency_dict[v] = 0

        # pattern_sequence = re.search(r'quence:(.*?)\Z', response, re.DOTALL).group(1)
        # pattern_sequence = re.search(r'(?i)pattern sequence.*?\n(.+?)\n*\n*\n', response, re.DOTALL).group(1)
        pattern_sequence = re.findall(r'\(V\d+, V\d+\):\s.*?(?=\n\n|\n|\Z)', response, re.DOTALL)
        pattern_dict = {}
        # lines = pattern_sequence.strip().split('\n')

        for line in pattern_sequence:
            key_value = line.split(':')
            key = tuple(key_value[0].strip()[1:-1].split(','))
            if all('V' in item for item in key):
                value = key_value[1].strip()
                pattern_dict[key] = value

        pattern_string = ""
        for key, value in pattern_dict.items():
            if len(key) == 2:
                frequency_dict[key[0]] += 1
                key_str = ", ".join(key)
                pattern_string += f"({key_str}): {value}\n"

        min_value = min(frequency_dict.values())
        candidate_ego = [key[-1] for key, value in frequency_dict.items() if value == min_value]
        func_scenario = "Initial actions: \n" + initial_string + "\n" + "Interactive pattern sequence: \n" + pattern_string
        func_scenario_dict = {"Road Structure": road_structure, "Initial Actions": initial_string, "Interactive Pattern Sequence": pattern_string}
        return func_scenario, func_scenario_dict, candidate_ego


if __name__ == "__main__":
    extractor = Extractor()
#     response = """Initial actions of vehicles:
# (V1, V2, V3): V1 was driving southbound in the second lane, V2 was stopped facing south in the third lane (a left hand turn lane) waiting to turn left, and V3 was also stopped in the third lane in front of V2 waiting to turn left.
#
# Interactive pattern sequence between pair of vehicles:
# (V1, V2): V1 drifted from the second lane to the third lane where V2 was stationary, waiting to turn left.
# (V2, V3): V2 was stationary behind V3, who was also stationary, waiting for the signal to turn left at the intersection.
# (V1, V3): The front of V1 contacted the back of V2, pushing V2 into the back of stationary V3."""
#     print(extractor.process_response(response))
    # workbook = openpyxl.load_workbook('../../data/accident_reports/straight_3_lane.xlsx')
    workbook = openpyxl.load_workbook('./data/accident_reports/cases.xlsx')
    sheet = workbook.active
    data_rows = sheet.iter_rows(min_row=20, max_row=20, values_only=True)
    for row in data_rows:
        report = row[1]
        print(report)
        extracted_data = extractor.extract(report)
        print(extracted_data)
